"""
Forward Rates Report Generator - Professional reporting for forward P&L analysis.
"""

from typing import Dict, List, Optional
import pandas as pd
import json
from datetime import datetime, timedelta
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import LineChart, Reference
from ..calculators.forward_pl_calculator import ForwardPLCalculator
from ..models.letter_of_credit import LetterOfCredit

logger = logging.getLogger(__name__)


class ForwardRatesReportGenerator:
    """
    Generates comprehensive reports for forward rates P&L analysis.
    """
    
    def __init__(self, forward_calculator: Optional[ForwardPLCalculator] = None):
        """Initialize the report generator."""
        self.forward_calculator = forward_calculator or ForwardPLCalculator()
    
    def generate_comprehensive_report(self, lc: LetterOfCredit, 
                                    base_currency: str = "INR") -> Dict[str, any]:
        """
        Generate comprehensive forward rates report.
        
        Args:
            lc: Letter of Credit instance
            base_currency: Base currency for calculations
        
        Returns:
            Complete report dictionary
        """
        try:
            logger.info(f"Generating comprehensive forward rates report for {lc.lc_id}")
            
            # Get forward P&L analysis
            forward_report = self.forward_calculator.generate_forward_pl_report(lc, base_currency)
            
            # Add executive summary
            executive_summary = self._create_executive_summary(forward_report)
            
            # Add risk assessment
            risk_assessment = self._create_risk_assessment(forward_report)
            
            # Add recommendations
            recommendations = self._create_recommendations(forward_report)
            
            # Compile comprehensive report
            comprehensive_report = {
                'report_metadata': {
                    'report_id': f"FWD_RPT_{lc.lc_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}",
                    'report_type': 'forward_rates_analysis',
                    'generated_at': datetime.now().isoformat(),
                    'lc_id': lc.lc_id,
                    'base_currency': base_currency,
                    'version': '1.0'
                },
                'executive_summary': executive_summary, 
                'forward_analysis': forward_report,
                'risk_assessment': risk_assessment,
                'recommendations': recommendations,
                'appendix': {
                    'methodology': self._get_methodology_notes(),
                    'data_sources': self._get_data_sources_info(),
                    'glossary': self._get_glossary()
                }
            }
            
            logger.info(f"Comprehensive report generated successfully")
            return comprehensive_report
            
        except Exception as e:
            logger.error(f"Error generating comprehensive report: {e}")
            return {}
    
    def export_to_excel(self, report: Dict[str, any], filename: str) -> str:
        """
        Export forward rates report to Excel with multiple sheets and charts.
        
        Args:
            report: Report dictionary
            filename: Output filename (without extension)
        
        Returns:
            Path to saved Excel file
        """
        try:
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create sheets
            self._create_summary_sheet(wb, report)
            self._create_daily_analysis_sheet(wb, report)
            self._create_scenarios_sheet(wb, report)
            self._create_charts_sheet(wb, report)
            self._create_recommendations_sheet(wb, report)
            
            # Save file
            filepath = f"{filename}.xlsx"
            wb.save(filepath)
            
            logger.info(f"Excel report saved: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {e}")
            return ""
    
    def export_to_json(self, report: Dict[str, any], filename: str) -> str:
        """Export report to JSON format."""
        try:
            filepath = f"{filename}.json"
            
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(report, f, indent=2, default=str, ensure_ascii=False)
            
            logger.info(f"JSON report saved: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error exporting to JSON: {e}")
            return ""
    
    def _create_executive_summary(self, forward_report: Dict) -> Dict[str, any]:
        """Create executive summary from forward report."""
        try:
            if not forward_report:
                return {}
            
            current_status = forward_report.get('current_status', {})
            analysis = forward_report.get('analysis', {})
            lc_details = forward_report.get('lc_details', {})
            
            # Determine overall status
            pl_percentage = current_status.get('pl_percentage', 0)
            if pl_percentage > 2:
                overall_status = "FAVORABLE"
                status_color = "green"
            elif pl_percentage < -2:
                overall_status = "UNFAVORABLE"
                status_color = "red"
            else:
                overall_status = "NEUTRAL"
                status_color = "yellow"
            
            # Calculate key metrics
            total_value = lc_details.get('total_value', 0)
            currency = lc_details.get('currency', 'USD')
            expected_value = current_status.get('expected_value_at_maturity', 0)
            pl_amount = current_status.get('unrealized_pl', 0)
            
            return {
                'overall_status': overall_status,
                'status_color': status_color,
                'key_metrics': {
                    'lc_value': f"{currency} {total_value:,.2f}",
                    'expected_maturity_value': f"₹{expected_value:,.2f}",
                    'forward_pl': f"₹{pl_amount:,.2f}",
                    'forward_pl_percentage': f"{pl_percentage:.2f}%",
                    'days_to_maturity': current_status.get('days_to_maturity', 0),
                    'forward_rate': f"₹{current_status.get('current_forward_rate', 0):.4f}"
                },
                'trend_analysis': {
                    'direction': analysis.get('forward_rate_trend', 'unknown'),
                    'volatility': f"{analysis.get('volatility', 0):.4f}",
                    'days_analyzed': analysis.get('total_days_tracked', 0)
                },
                'summary_text': self._generate_summary_text(overall_status, pl_amount, pl_percentage)
            }
            
        except Exception as e:
            logger.error(f"Error creating executive summary: {e}")
            return {}
    
    def _create_risk_assessment(self, forward_report: Dict) -> Dict[str, any]:
        """Create risk assessment from forward report."""
        try:
            analysis = forward_report.get('analysis', {})
            current_status = forward_report.get('current_status', {})
            
            # Risk factors
            pl_percentage = abs(current_status.get('pl_percentage', 0))
            volatility = analysis.get('volatility', 0)
            days_remaining = current_status.get('days_to_maturity', 0)
            
            # Calculate risk scores
            pl_risk_score = min(10, pl_percentage * 2)  # Scale to 10
            volatility_risk_score = min(10, volatility * 1000)  # Scale to 10
            time_risk_score = max(1, min(10, 10 - (days_remaining / 10)))  # More risk with less time
            
            overall_risk_score = (pl_risk_score + volatility_risk_score + time_risk_score) / 3
            
            # Risk level
            if overall_risk_score <= 3:
                risk_level = "LOW"
                risk_color = "green"
            elif overall_risk_score <= 6:
                risk_level = "MEDIUM"
                risk_color = "yellow"
            else:
                risk_level = "HIGH"
                risk_color = "red"
            
            return {
                'overall_risk_level': risk_level,
                'overall_risk_score': round(overall_risk_score, 2),
                'risk_color': risk_color,
                'risk_factors': {
                    'pl_risk': {
                        'score': round(pl_risk_score, 2),
                        'description': f"P&L deviation of {pl_percentage:.2f}%"
                    },
                    'volatility_risk': {
                        'score': round(volatility_risk_score, 2),
                        'description': f"Forward rate volatility of {volatility:.4f}"
                    },
                    'time_risk': {
                        'score': round(time_risk_score, 2),
                        'description': f"{days_remaining} days remaining to maturity"
                    }
                },
                'risk_mitigation': self._get_risk_mitigation_strategies(risk_level, current_status)
            }
            
        except Exception as e:
            logger.error(f"Error creating risk assessment: {e}")
            return {}
    
    def _create_recommendations(self, forward_report: Dict) -> Dict[str, any]:
        """Create actionable recommendations."""
        try:
            current_status = forward_report.get('current_status', {})
            analysis = forward_report.get('analysis', {})
            exit_scenarios = forward_report.get('exit_scenarios', [])
            hold_scenario = forward_report.get('hold_to_maturity_scenario', {})
            
            recommendations = []
            
            # Based on current P&L
            pl_percentage = current_status.get('pl_percentage', 0)
            
            if pl_percentage > 3:
                recommendations.append({
                    'priority': 'HIGH',
                    'action': 'CONSIDER_LOCKING_GAINS',
                    'description': 'Forward rates are favorable. Consider hedging to lock in gains.',
                    'rationale': f'Current forward P&L of {pl_percentage:.2f}% represents significant gain'
                })
            elif pl_percentage < -3:
                recommendations.append({
                    'priority': 'HIGH', 
                    'action': 'CONSIDER_EXIT_OR_HEDGE',
                    'description': 'Forward rates are unfavorable. Consider early exit or hedging.',
                    'rationale': f'Current forward P&L of {pl_percentage:.2f}% represents significant loss'
                })
            
            # Based on trend
            trend = analysis.get('forward_rate_trend', '')
            if trend == 'strengthening':
                recommendations.append({
                    'priority': 'MEDIUM',
                    'action': 'MONITOR_FOR_PEAK',
                    'description': 'Forward rates are strengthening. Monitor for optimal exit point.',
                    'rationale': 'Upward trend may continue, but be ready to act at peak'
                })
            elif trend == 'weakening':
                recommendations.append({
                    'priority': 'MEDIUM',
                    'action': 'PREPARE_HEDGING_STRATEGY',
                    'description': 'Forward rates are weakening. Prepare hedging strategy.',
                    'rationale': 'Downward trend may continue, prepare defensive measures'
                })
            
            # Based on volatility
            volatility = analysis.get('volatility', 0)
            if volatility > 0.02:  # High volatility threshold
                recommendations.append({
                    'priority': 'MEDIUM',
                    'action': 'INCREASE_MONITORING',
                    'description': 'High volatility detected. Increase monitoring frequency.',
                    'rationale': f'Volatility of {volatility:.4f} indicates unstable market conditions'
                })
            
            # Scenario-based recommendations
            if exit_scenarios:
                best_exit = max(exit_scenarios, key=lambda x: x.get('pl_percentage', -999))
                if best_exit.get('pl_percentage', 0) > pl_percentage:
                    recommendations.append({
                        'priority': 'LOW',
                        'action': 'CONSIDER_EARLY_EXIT',
                        'description': f'Early exit on {best_exit.get("exit_date")} may be more favorable.',
                        'rationale': f'Exit scenario shows {best_exit.get("pl_percentage", 0):.2f}% vs current {pl_percentage:.2f}%'
                    })
            
            return {
                'recommendations': recommendations,
                'next_review_date': self._calculate_next_review_date(),
                'key_monitoring_points': [
                    'Daily forward rate changes > 0.5%',
                    'P&L percentage crossing ±2% threshold',
                    'Volatility spike above normal levels',
                    'Major economic announcements affecting USD/INR'
                ]
            }
            
        except Exception as e:
            logger.error(f"Error creating recommendations: {e}")
            return {}
    
    def _create_summary_sheet(self, wb: Workbook, report: Dict):
        """Create executive summary sheet in Excel."""
        ws = wb.create_sheet("Executive Summary")
        
        # Headers and styling
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        
        # Title
        ws['A1'] = "Forward Rates P&L Analysis - Executive Summary"
        ws['A1'].font = Font(bold=True, size=16)
        
        # Summary data
        summary = report.get('executive_summary', {})
        row = 3
        
        ws[f'A{row}'] = "Overall Status:"
        ws[f'B{row}'] = summary.get('overall_status', 'N/A')
        ws[f'B{row}'].font = Font(bold=True)
        row += 2
        
        # Key metrics
        metrics = summary.get('key_metrics', {})
        ws[f'A{row}'] = "Key Metrics:"
        ws[f'A{row}'].font = subheader_font
        row += 1
        
        for key, value in metrics.items():
            ws[f'A{row}'] = key.replace('_', ' ').title() + ":"
            ws[f'B{row}'] = value
            row += 1
    
    def _create_daily_analysis_sheet(self, wb: Workbook, report: Dict):
        """Create daily analysis sheet in Excel."""
        ws = wb.create_sheet("Daily Analysis")
        
        # Headers
        headers = ['Date', 'Forward Rate', 'Expected Value', 'P&L Amount', 'P&L %', 'Daily Change']
        for i, header in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=header).font = Font(bold=True)
        
        # Data
        daily_pl = report.get('forward_analysis', {}).get('daily_forward_pl', {})
        row = 2
        
        for date, data in sorted(daily_pl.items()):
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=data.get('forward_rate', 0))
            ws.cell(row=row, column=3, value=data.get('expected_value_at_maturity', 0))
            ws.cell(row=row, column=4, value=data.get('unrealized_pl', 0))
            ws.cell(row=row, column=5, value=data.get('pl_percentage', 0))
            ws.cell(row=row, column=6, value=data.get('daily_change', 0))
            row += 1
    
    def _create_scenarios_sheet(self, wb: Workbook, report: Dict):
        """Create scenarios analysis sheet in Excel."""
        ws = wb.create_sheet("Scenarios")
        
        # Headers
        ws['A1'] = "Scenario Analysis"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        
        # Hold to maturity scenario
        hold_scenario = report.get('forward_analysis', {}).get('hold_to_maturity_scenario', {})
        if hold_scenario:
            ws[f'A{row}'] = "Hold to Maturity Scenario:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            ws[f'A{row}'] = "Expected P&L:"
            ws[f'B{row}'] = f"₹{hold_scenario.get('unrealized_pl', 0):,.2f} ({hold_scenario.get('pl_percentage', 0):.2f}%)"
            row += 2
        
        # Exit scenarios
        exit_scenarios = report.get('forward_analysis', {}).get('exit_scenarios', [])
        if exit_scenarios:
            ws[f'A{row}'] = "Early Exit Scenarios:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            
            headers = ['Exit Date', 'Days Held', 'P&L Amount', 'P&L %']
            for i, header in enumerate(headers):
                ws.cell(row=row, column=i+1, value=header).font = Font(bold=True)
            row += 1
            
            for scenario in exit_scenarios:
                ws.cell(row=row, column=1, value=scenario.get('exit_date', ''))
                ws.cell(row=row, column=2, value=scenario.get('days_held', 0))
                ws.cell(row=row, column=3, value=scenario.get('unrealized_pl', 0))
                ws.cell(row=row, column=4, value=scenario.get('pl_percentage', 0))
                row += 1
    
    def _create_charts_sheet(self, wb: Workbook, report: Dict):
        """Create charts sheet in Excel."""
        ws = wb.create_sheet("Charts")
        ws['A1'] = "Forward Rates Charts"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Note: Excel chart creation would require more complex openpyxl chart setup
        ws['A3'] = "Charts will be added in future version"
    
    def _create_recommendations_sheet(self, wb: Workbook, report: Dict):
        """Create recommendations sheet in Excel."""
        ws = wb.create_sheet("Recommendations")
        
        ws['A1'] = "Recommendations & Action Items"
        ws['A1'].font = Font(bold=True, size=14)
        
        recommendations = report.get('recommendations', {}).get('recommendations', [])
        row = 3
        
        headers = ['Priority', 'Action', 'Description', 'Rationale']
        for i, header in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=header).font = Font(bold=True)
        row += 1
        
        for rec in recommendations:
            ws.cell(row=row, column=1, value=rec.get('priority', ''))
            ws.cell(row=row, column=2, value=rec.get('action', ''))
            ws.cell(row=row, column=3, value=rec.get('description', ''))
            ws.cell(row=row, column=4, value=rec.get('rationale', ''))
            row += 1
    
    def _get_methodology_notes(self) -> List[str]:
        """Get methodology notes for the report."""
        return [
            "Forward rates calculated using Interest Rate Parity theory",
            "Volatility estimated from historical exchange rate movements",
            "Risk assessment based on P&L deviation and market volatility",
            "Scenarios modeled using current market conditions",
            "Recommendations based on quantitative analysis and market trends"
        ]
    
    def _get_data_sources_info(self) -> List[str]:
        """Get data sources information."""
        return [
            "Yahoo Finance for historical exchange rates",
            "Calculated forward rates using IRP methodology",
            "Market volatility from historical price movements",
            "Interest rates estimated from market standards"
        ]
    
    def _get_glossary(self) -> Dict[str, str]:
        """Get glossary of terms."""
        return {
            "Forward Rate": "Expected exchange rate for a future date",
            "Forward P&L": "Profit/Loss based on forward rate expectations",
            "Interest Rate Parity": "Theory linking exchange rates and interest rates",
            "Volatility": "Measure of exchange rate variability",
            "VaR": "Value at Risk - potential loss at given confidence level",
            "Unrealized P&L": "Paper profit/loss that hasn't been realized yet"
        }
    
    def _generate_summary_text(self, status: str, pl_amount: float, pl_percentage: float) -> str:
        """Generate executive summary text."""
        if status == "FAVORABLE":
            return f"Forward market expectations are favorable with a projected gain of ₹{pl_amount:,.2f} ({pl_percentage:.2f}%). Consider strategies to lock in these gains."
        elif status == "UNFAVORABLE":
            return f"Forward market expectations are unfavorable with a projected loss of ₹{abs(pl_amount):,.2f} ({abs(pl_percentage):.2f}%). Consider hedging or early exit strategies."
        else:
            return f"Forward market expectations are neutral with minimal impact of ₹{pl_amount:,.2f} ({pl_percentage:.2f}%). Continue monitoring for trend changes."
    
    def _get_risk_mitigation_strategies(self, risk_level: str, current_status: Dict) -> List[str]:
        """Get risk mitigation strategies based on risk level."""
        strategies = []
        
        if risk_level == "HIGH":
            strategies.extend([
                "Consider immediate hedging to limit losses",
                "Evaluate early exit options",
                "Increase monitoring frequency to daily",
                "Set stop-loss levels"
            ])
        elif risk_level == "MEDIUM":
            strategies.extend([
                "Monitor position closely",
                "Prepare hedging strategy",
                "Review position weekly",
                "Set alert thresholds"
            ])
        else:  # LOW
            strategies.extend([
                "Continue regular monitoring",
                "Review position monthly",
                "Maintain current strategy"
            ])
        
        return strategies
    
    def _calculate_next_review_date(self) -> str:
        """Calculate next review date."""
        next_review = datetime.now() + timedelta(days=7)  # Weekly review
        return next_review.strftime("%Y-%m-%d")
